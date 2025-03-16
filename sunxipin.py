from argparse import ArgumentParser
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from natsort import natsorted
import pandas as pd

parser = ArgumentParser()
parser.add_argument("input", help="Input file")
parser.add_argument("output", help="Output file")


def load(file) -> DataFrame:
    """Load a file into a DataFrame"""
    df = pd.read_excel(file)
    df.iloc[:, 0] = df.iloc[:, 0].ffill()
    return df


def table_columns(df: DataFrame, like: str) -> str:
    reslut = []
    for col in df.columns:
        if like in col:
            reslut.append(col)
    if len(reslut) == 1:
        return reslut[0]
    else:
        raise Exception("Multiple columns found")


def functions(df: DataFrame) -> list:
    funcs = set()
    function_table = df.filter(like="Function")
    for line in function_table.values:
        for elem in line:
            if elem and not pd.isna(elem):
                funcs.add(elem)
    return natsorted(funcs)


def get_pins(df: DataFrame, function: str) -> list:
    mask = df.map(lambda x: function == x)
    rows = df.loc[mask.any(axis=1)]
    col = table_columns(rows, "Pin")
    return rows[col].to_list()

# 不能随便改
def write_function_sheet(sheet, df: DataFrame):
    def real_addr(cell) -> str:
        row = cell.row
        column = cell.column_letter
        return f"${column}${row}"

    # write functions
    sheet.append(["Function", "Count", "Pins"])
    maxlen = 0
    for func in functions(df):
        pins = get_pins(df, func)
        sheet.append([func, len(pins)] + pins)
        maxlen = max(maxlen, len(pins))
    if maxlen:
        sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + maxlen)

FLITER = ["LCD", "UART", "SDC0", "SPI0", "SDC2", "RMII", "TWI"]

# 添加一个演示代码
def write_helper_sheet(sheet, title, df):
    sheet.append(["Function", "Pin"])
    target_cell = "A2" 
    line = f"MATCH({target_cell}, '{title}'!$A:$A, 0)"
    count = f"INDEX('{title}'!$B:$B, {line})"
    range = f"OFFSET('{title}'!$A$1, {line} - 1, 2, 1, {count})"
    dv = DataValidation(type="list", formula1=f"={range}")
    sheet.add_data_validation(dv)
    # sheet.append(["LCD-D2", f"={line}", f"={count}", f"={range}"]) # 测试公式]
    for func in functions(df):
        if all(fliter not in func for fliter in FLITER):
            continue
        pins = get_pins(df, func)
        if len(pins) == 1:
            sheet.append([func, pins[0]])
        else:
            sheet.append([func, None])
        dv.add(sheet[f"B{sheet.max_row}"])


if __name__ == "__main__":
    args = parser.parse_args()
    df = load(args.input)
    excel = Workbook()
    excel.remove(excel.active)  # remove default sheet
    excel.create_sheet("Helper")
    excel.create_sheet("Pins")
    write_function_sheet(excel["Pins"], df)
    write_helper_sheet(excel["Helper"], excel["Pins"].title, df)
    excel.save(args.output)
